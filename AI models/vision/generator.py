"""
HireMind - Report Generator  (v3.1)
====================================
Changes vs v3.0:
  - generate() now writes TWO outputs:
      1. <session_id>_frames.jsonl  — one line per frame, target format:
         {question, timestamp, age, gender, emotion, face_count, speech,
          attention_status, head_pitch, head_yaw, eye_aspect_ratio, mouth_ratio}
      2. HireMind_Report_<id>.json + .xlsx — full summary report (unchanged)
  - _write_frames_jsonl() reads history and re-emits the frame log in the
    exact target shape (age/gender/emotion filled from DeepFace when available).
"""

import json
import os
import time
from datetime import datetime
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _XLSX_AVAILABLE = True
except ImportError:
    _XLSX_AVAILABLE = False


# ─────────────────────────────────────────────────────────────────
# Colour palette for Excel
# ─────────────────────────────────────────────────────────────────
CLR_HEADER   = "1A3A5C"
CLR_ACCENT   = "E07B00"
CLR_GOOD     = "1B7F4B"
CLR_WARN     = "C0392B"
CLR_NEUTRAL  = "2C3E50"
CLR_ROW_ALT  = "EAF0FB"
CLR_WHITE    = "FFFFFF"
CLR_LABEL    = "2E4053"


def _thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


class ReportGenerator:
    """Generates the post-interview frame JSONL + summary JSON + Excel report."""

    def __init__(self, output_dir: str = "."):
        self.output_dir = output_dir
        os.makedirs(output_dir, exist_ok=True)

    # ──────────────────────────────────────────────────
    # Public API
    # ──────────────────────────────────────────────────

    def generate(self, session_report: dict, history: list, session_id: str = "") -> dict:
        """
        Generate all outputs with a unique, human-readable session ID.
        Files:
          1. <session_id>_frames.jsonl  — one line per frame (full target format)
          2. HireMind_Report_<session_id>.json
          3. HireMind_Report_<session_id>.xlsx  — Summary + Frame Log + Q&A + Timeline
        """
        # ── Unique session ID: date_time_random ─────────────────────────────────
        import random, string
        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        rand     = "".join(random.choices(string.ascii_lowercase + string.digits, k=6))
        uid      = session_id if session_id else f"{ts}_{rand}"
        base_name = f"HireMind_Report_{uid}"

        frames_path = self._write_frames_jsonl(history, uid, base_name)
        json_path   = self._write_json(session_report, history, base_name)
        excel_path  = self._write_excel(session_report, history, base_name)

        return {
            "session_id":  uid,
            "frames_path": frames_path,
            "json_path":   json_path,
            "excel_path":  excel_path,
            "generated_at": datetime.now().isoformat(),
        }

    # ──────────────────────────────────────────────────
    # Frame JSONL  ← TARGET FORMAT
    # ──────────────────────────────────────────────────

    def _write_frames_jsonl(self, history: list, session_id: str, base: str) -> str:
        """
        Write one JSON line per history entry in the exact target format:
        {question, timestamp, age, gender, emotion, face_count,
         attention_status, eye_contact_pct, head_pitch, head_yaw, head_roll,
         eye_aspect_ratio, mouth_ratio, eye_movement_status,
         motion_score, motion_state, stress_level, stress_score,
         deception_cue_level, speech,
         confidence_score, fluency_score, hesitation_count,
         speech_rate_wpm, filler_count, word_count,
         speaking_duration, energy_variance,
         star_score, overall_score, communication_score, engagement_score,
         situation, task, action, result_dim, grade}
        """
        path = os.path.join(self.output_dir, f"{base}_frames.jsonl")
        q_counter = 0
        prev_q    = None

        with open(path, "w", encoding="utf-8") as fh:
            for entry in history:
                video    = entry.get("video",    {})
                audio    = entry.get("audio",    {})
                feedback = entry.get("feedback", {})
                scores   = entry.get("scores",   {})

                face    = video.get("face",    {})
                posture = video.get("posture", {})
                eye     = video.get("eye",     {})
                emotion = video.get("emotion", {})
                stress  = video.get("stress",  {})

                # ── Question number ───────────────────────────────────────────
                q_text = (feedback.get("next_question") or
                          entry.get("question") or prev_q or "")
                if q_text and q_text != prev_q:
                    q_counter += 1
                    prev_q = q_text
                q_num = str(q_counter) if q_counter else "1"

                # ── Emotion ───────────────────────────────────────────────────
                df      = emotion.get("deepface") or {}
                mp_emo  = emotion.get("mediapipe") or emotion
                emotion_label = (df.get("dominant_emotion")
                                 or mp_emo.get("label") or "N/A")
                age    = df.get("age",    "N/A")
                gender = df.get("gender", "N/A")
                if isinstance(gender, dict):
                    gender = max(gender, key=gender.get)

                # ── Head pose ─────────────────────────────────────────────────
                pose       = face.get("head_pose", {})
                head_pitch = round(pose.get("pitch", 0.0), 2) if face.get("detected") else None
                head_yaw   = round(pose.get("yaw",   0.0), 2) if face.get("detected") else None
                head_roll  = round(pose.get("roll",  0.0), 2) if face.get("detected") else None

                # ── Eye / mouth ───────────────────────────────────────────────
                ear_val   = video.get("_eye_aspect_ratio", None)
                mouth_val = video.get("_mouth_ratio",      None)
                if ear_val is not None:   ear_val   = round(ear_val,   3)
                if mouth_val is not None: mouth_val = round(mouth_val, 3)

                # ── Eye contact ───────────────────────────────────────────────
                contact_ratio = eye.get("contact_ratio", 0.0)
                eye_contact_pct = round(contact_ratio * 100, 1)
                looking_away  = eye.get("looking_away", False)
                gaze_dir      = eye.get("gaze_direction", "center")
                eye_mov_status = "away" if looking_away else "center"

                # ── Stress ────────────────────────────────────────────────────
                stress_level  = stress.get("level", "unknown").lower() if stress else "unknown"
                stress_score  = stress.get("score", None) if stress else None

                # ── Motion (from head instability proxy) ──────────────────────
                tension = face.get("facial_tension", 0.0)
                motion_score = round(tension, 3)
                if tension < 0.2:   motion_state = "still"
                elif tension < 0.5: motion_state = "slight"
                else:               motion_state = "active"

                # ── Speech / transcript ───────────────────────────────────────
                transcript  = audio.get("speech_transcript", "")
                analysis    = audio.get("analysis", {}) or {}
                wpm         = analysis.get("words_per_minute", 0)
                filler_cnt  = analysis.get("filler_count",     0)
                clarity_sc  = analysis.get("clarity_score",    0.0)
                words       = len(transcript.split()) if transcript else 0
                speaking_dur= round(words / max(wpm, 1) * 60, 1) if wpm else 0.0

                # ── Scores from coach ─────────────────────────────────────────
                conf_sc     = scores.get("confidence",    0.0) or 0.0
                eye_sc      = scores.get("eye_contact",   0.0) or 0.0
                posture_sc  = scores.get("posture",       0.0) or 0.0
                ans_sc      = scores.get("answer_quality",0.0) or 0.0
                overall_sc  = scores.get("overall",       0.0) or 0.0
                comm_sc     = round((conf_sc + clarity_sc) / 2, 1)
                engage_sc   = round((eye_sc  + posture_sc) / 2, 1)

                # ── STAR dimensions from text_result if present ───────────────
                text_r   = entry.get("text_result", {}) or {}
                star_sig = text_r.get("star_signals", [])
                situation_sc = 1.0 if "situation" in star_sig else 0.0
                task_sc      = 1.0 if "task"      in star_sig else 0.0
                action_sc    = 1.0 if "action"    in star_sig else 0.0
                result_sc    = 1.0 if "result"    in star_sig else 0.0
                star_total   = round((situation_sc + task_sc + action_sc + result_sc) / 4 * 10, 1)

                # ── Grade ─────────────────────────────────────────────────────
                grade = "N/A"
                if overall_sc >= 8.5:   grade = "A"
                elif overall_sc >= 7.0: grade = "B"
                elif overall_sc >= 5.5: grade = "C"
                elif overall_sc >= 4.0: grade = "D"
                elif overall_sc > 0:    grade = "F"

                # ── Attention status ──────────────────────────────────────────
                slouching = posture.get("slouching", False)
                if transcript:              attention = "speaking"
                elif ear_val and ear_val < 0.18: attention = "blink/eyes low"
                elif looking_away:          attention = "distracted"
                elif slouching:             attention = "slouching"
                else:                       attention = "focused"

                record = {
                    "question":            q_num,
                    "timestamp":           entry.get("_ts", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
                    "age":                 age,
                    "gender":              gender,
                    "emotion":             emotion_label,
                    "face_count":          1 if face.get("detected") else 0,
                    "attention_status":    attention,
                    "eye_contact_pct":     eye_contact_pct,
                    "head_pitch":          head_pitch,
                    "head_yaw":            head_yaw,
                    "head_roll":           head_roll,
                    "eye_aspect_ratio":    ear_val,
                    "mouth_ratio":         mouth_val,
                    "eye_movement_status": eye_mov_status,
                    "motion_score":        motion_score,
                    "motion_state":        motion_state,
                    "stress_level":        stress_level,
                    "stress_score":        stress_score,
                    "deception_cue_level": "inconclusive",
                    "speech":              transcript,
                    "confidence_score":    round(conf_sc,  1),
                    "fluency_score":       round(clarity_sc, 1),
                    "hesitation_count":    filler_cnt,
                    "speech_rate_wpm":     wpm,
                    "filler_count":        filler_cnt,
                    "word_count":          words,
                    "speaking_duration":   speaking_dur,
                    "energy_variance":     round(tension * 10, 2),
                    "star_score":          star_total,
                    "overall_score":       round(overall_sc, 1),
                    "communication_score": comm_sc,
                    "engagement_score":    engage_sc,
                    "situation":           situation_sc,
                    "task":                task_sc,
                    "action":              action_sc,
                    "result_dim":          result_sc,
                    "grade":               grade,
                }
                fh.write(json.dumps(record, ensure_ascii=False) + "\n")

        return path

    # ──────────────────────────────────────────────────
    # Summary JSON
    # ──────────────────────────────────────────────────

    def _write_json(self, session_report: dict, history: list, base: str) -> str:
        path = os.path.join(self.output_dir, f"{base}.json")

        q_breakdown = self._question_breakdown(history)

        report = {
            "hiremind_report": {
                "version":    "1.0",
                "session_id": base,
                "generated":  datetime.now().isoformat(),
                "summary":    session_report,
                "question_by_question": q_breakdown,
                "timeline": [
                    {
                        "t":      round(h["timestamp"], 1),
                        "scores": h.get("scores", {}),
                    }
                    for h in history[::5]
                ],
                "warnings": self._extract_warnings(session_report),
            }
        }

        with open(path, "w", encoding="utf-8") as f:
            json.dump(report, f, indent=2, ensure_ascii=False)

        return path

    # ──────────────────────────────────────────────────
    # Excel
    # ──────────────────────────────────────────────────

    def _write_excel(self, session_report: dict, history: list, base: str) -> str:
        path = os.path.join(self.output_dir, f"{base}.xlsx")

        if not _XLSX_AVAILABLE:
            print("[ReportGenerator] openpyxl not installed — skipping Excel output.")
            return ""

        wb = openpyxl.Workbook()

        self._sheet_summary(wb, session_report)
        self._sheet_frames(wb, history)
        self._sheet_qa(wb, history)
        self._sheet_timeline(wb, history)

        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(path)
        return path

    # ── Summary sheet ────────────────────────────────

    def _sheet_summary(self, wb, report: dict):
        ws = wb.create_sheet("Summary", 0)
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 16
        ws.column_dimensions["C"].width = 18

        def hdr(row, text, color=CLR_HEADER):
            c = ws.cell(row=row, column=1, value=text)
            c.font      = Font(bold=True, color=CLR_WHITE, size=11, name="Arial")
            c.fill      = PatternFill("solid", fgColor=color)
            c.alignment = Alignment(horizontal="center")
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
            c.border = _thin_border()

        def row2(r, label, value, note=""):
            lc = ws.cell(row=r, column=1, value=label)
            vc = ws.cell(row=r, column=2, value=value)
            nc = ws.cell(row=r, column=3, value=note)
            lc.font      = Font(bold=True, color=CLR_LABEL, name="Arial", size=10)
            vc.alignment = Alignment(horizontal="center")
            vc.font      = Font(name="Arial", size=10)
            nc.font      = Font(name="Arial", size=10)
            for c in [lc, vc, nc]:
                c.border = _thin_border()
            # colour-code the score cell
            try:
                v = float(value)
                if v >= 7.5:
                    vc.fill = PatternFill("solid", fgColor="D5F5E3")
                    vc.font = Font(bold=True, color=CLR_GOOD, name="Arial", size=10)
                elif v < 5.5:
                    vc.fill = PatternFill("solid", fgColor="FADBD8")
                    vc.font = Font(bold=True, color=CLR_WARN, name="Arial", size=10)
            except (TypeError, ValueError):
                pass

        r = 1
        # ── Title ────────────────────────────────────────────────────────────
        hdr(r, "HireMind — Session Summary"); r += 1
        ws.cell(row=r, column=1, value="Generated").font = Font(italic=True, color="888888", name="Arial")
        ws.cell(row=r, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))
        r += 2

        # ── Overall Performance ───────────────────────────────────────────────
        hdr(r, "Overall Performance", CLR_ACCENT); r += 1
        avg = report.get("average_scores", {})
        score_rows = [
            ("overall",       "Overall"),
            ("confidence",    "Confidence"),
            ("clarity",       "Clarity"),
            ("posture",       "Posture"),
            ("eye_contact",   "Eye Contact"),
            ("answer_quality","Answer Quality"),
        ]
        for key, label in score_rows:
            val  = avg.get(key, "—")
            note = ""
            if isinstance(val, (int, float)):
                if   val >= 7.5: note = "★ Strong"
                elif val < 5.5:  note = "⚠ Improve"
            row2(r, label, val, note); r += 1

        r += 1
        # ── Session Info ──────────────────────────────────────────────────────
        hdr(r, "Session Info"); r += 1
        dur   = report.get("session_duration_min", "—")
        frames= report.get("total_frames", "—")
        trend = report.get("trend", "—")
        strengths    = ", ".join(report.get("strengths",             [])) or "—"
        improvements = ", ".join(report.get("areas_for_improvement", [])) or "—"
        chronic      = ", ".join(report.get("chronic_issues",        [])) or "None"

        row2(r, "Duration (min)",  dur);        r += 1
        row2(r, "Total Frames",    frames);     r += 1
        row2(r, "Trend",           trend);      r += 1
        row2(r, "Key Strengths",   strengths);  r += 1
        row2(r, "Areas to Improve",improvements); r += 1
        row2(r, "Chronic Issues",  chronic)

    # ── Frame-by-frame sheet (target format) ─────────

    def _sheet_frames(self, wb, history: list):
        ws = wb.create_sheet("Frame Log", 1)

        headers = [
            "Question", "Timestamp", "Age", "Gender", "Emotion",
            "Face Count", "Speech", "Attention Status",
            "Head Pitch", "Head Yaw", "Eye Aspect Ratio", "Mouth Ratio",
        ]
        col_widths = [30, 20, 8, 10, 14, 10, 40, 18, 12, 12, 16, 12]

        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font      = Font(bold=True, color=CLR_WHITE, size=10)
            c.fill      = PatternFill("solid", fgColor=CLR_HEADER)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border    = _thin_border()
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.row_dimensions[1].height = 28

        for ri, entry in enumerate(history, 2):
            video   = entry.get("video",    {})
            audio   = entry.get("audio",    {})
            feedback= entry.get("feedback", {})

            face    = video.get("face",    {})
            posture = video.get("posture", {})
            eye     = video.get("eye",     {})
            emotion = video.get("emotion", {})

            df     = emotion.get("deepface") or {}
            mp_emo = emotion.get("mediapipe") or emotion
            emotion_label = df.get("dominant_emotion") or mp_emo.get("label") or "N/A"

            age    = df.get("age",    "N/A")
            gender = df.get("gender", "N/A")
            if isinstance(gender, dict):
                gender = max(gender, key=gender.get)

            pose       = face.get("head_pose", {})
            head_pitch = round(pose.get("pitch", 0.0), 2)
            head_yaw   = round(pose.get("yaw",   0.0), 2)
            ear_val    = round(video.get("_eye_aspect_ratio", 0.0), 3)
            mouth_val  = round(video.get("_mouth_ratio",      0.0), 3)

            transcript   = audio.get("speech_transcript", "")
            looking_away = eye.get("looking_away", False)
            slouching    = posture.get("slouching", False)

            if transcript:
                attention = "speaking"
            elif 0 < ear_val < 0.18:
                attention = "blink/eyes low"
            elif looking_away:
                attention = "distracted"
            elif slouching:
                attention = "slouching"
            else:
                attention = "focused"

            question = feedback.get("next_question", "")

            fill = PatternFill("solid", fgColor=CLR_ROW_ALT if ri % 2 == 0 else CLR_WHITE)
            row_vals = [
                question, datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                age, gender, emotion_label,
                1 if face.get("detected") else 0,
                transcript, attention,
                head_pitch, head_yaw, ear_val, mouth_val,
            ]
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill      = fill
                c.border    = _thin_border()
                c.alignment = Alignment(wrap_text=True, vertical="top")

    # ── Q&A sheet ────────────────────────────────────

    def _sheet_qa(self, wb, history: list):
        ws = wb.create_sheet("Q&A Analysis")

        headers    = ["Question", "Answer (transcript)", "Speech Score",
                      "Emotion Score", "Eye Contact Score", "Answer Quality", "Final Feedback"]
        col_widths = [35, 45, 14, 14, 16, 14, 40]

        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font      = Font(bold=True, color=CLR_WHITE, size=10)
            c.fill      = PatternFill("solid", fgColor=CLR_HEADER)
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            c.border    = _thin_border()
            ws.column_dimensions[get_column_letter(ci)].width = w

        ws.row_dimensions[1].height = 28
        qa_entries = self._question_breakdown(history)

        for ri, entry in enumerate(qa_entries, 2):
            fill   = PatternFill("solid", fgColor=CLR_ROW_ALT if ri % 2 == 0 else CLR_WHITE)
            scores = entry.get("avg_scores", {})

            def _score(k):
                v = scores.get(k)
                return round(v, 1) if isinstance(v, (int, float)) else "—"

            row_vals = [
                entry.get("question", "—"),
                entry.get("transcript_sample", "—"),
                _score("clarity"), _score("confidence"),
                _score("eye_contact"), _score("answer_quality"),
                entry.get("feedback_summary", "—"),
            ]
            for ci, val in enumerate(row_vals, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill      = fill
                c.border    = _thin_border()
                c.alignment = Alignment(wrap_text=True, vertical="top")
                if ci in (3, 4, 5, 6) and isinstance(val, float):
                    if val >= 7.5:
                        c.font = Font(bold=True, color=CLR_GOOD)
                    elif val < 5.5:
                        c.font = Font(bold=True, color=CLR_WARN)
            ws.row_dimensions[ri].height = 55

    # ── Timeline sheet ───────────────────────────────

    def _sheet_timeline(self, wb, history: list):
        ws = wb.create_sheet("Score Timeline")

        headers = ["Time (s)", "Confidence", "Clarity", "Posture",
                   "Eye Contact", "Answer Quality", "Overall"]
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color=CLR_WHITE)
            c.fill = PatternFill("solid", fgColor=CLR_NEUTRAL)
            ws.column_dimensions[get_column_letter(ci)].width = 14

        for ri, entry in enumerate(history[::3], 2):
            scores = entry.get("scores", {})
            row_data = [
                round(entry.get("timestamp", 0), 1),
                scores.get("confidence"), scores.get("clarity"),
                scores.get("posture"),    scores.get("eye_contact"),
                scores.get("answer_quality"), scores.get("overall"),
            ]
            fill = PatternFill("solid", fgColor=CLR_ROW_ALT if ri % 2 == 0 else CLR_WHITE)
            for ci, val in enumerate(row_data, 1):
                c = ws.cell(row=ri, column=ci, value=val)
                c.fill   = fill
                c.border = _thin_border()

    # ──────────────────────────────────────────────────
    # Helpers
    # ──────────────────────────────────────────────────

    def _question_breakdown(self, history: list) -> list:
        entries   = []
        current_q = None
        block     = []

        for h in history:
            feedback = h.get("feedback", {})
            q = feedback.get("next_question")
            if q and q != current_q:
                if block and current_q:
                    entries.append(self._summarise_block(current_q, block))
                current_q = q
                block = []
            block.append(h)

        if block and current_q:
            entries.append(self._summarise_block(current_q, block))

        return entries

    def _summarise_block(self, question: str, block: list) -> dict:
        import numpy as np
        all_scores = [h.get("scores", {}) for h in block]
        keys = ["confidence", "clarity", "posture", "eye_contact", "answer_quality", "overall"]
        avg_scores = {}
        for k in keys:
            vals = [s.get(k) for s in all_scores if s.get(k) is not None]
            avg_scores[k] = round(float(np.mean(vals)), 1) if vals else None

        transcripts = []
        for h in block:
            audio = h.get("audio", {})
            t = audio.get("speech_transcript", "") or audio.get("transcript", "")
            if t:
                transcripts.append(t)

        tips = []
        for h in block:
            fb = h.get("feedback", {})
            tips.extend(fb.get("tips", []))
        feedback_summary = "; ".join(dict.fromkeys(tips))[:200] if tips else "Good performance"

        return {
            "question":          question,
            "transcript_sample": " ".join(transcripts)[:300] if transcripts else "—",
            "avg_scores":        avg_scores,
            "feedback_summary":  feedback_summary,
        }

    def _extract_warnings(self, report: dict) -> list:
        warnings = []
        avg = report.get("average_scores", {})
        for k, v in avg.items():
            if isinstance(v, float) and v < 5.0:
                warnings.append(f"Low {k.replace('_', ' ')}: {v}/10")
        for issue in report.get("chronic_issues", []):
            warnings.append(f"Chronic issue: {issue.replace('_', ' ')}")
        if report.get("trend") == "declining":
            warnings.append("Performance trend declined over the session.")
        return warnings